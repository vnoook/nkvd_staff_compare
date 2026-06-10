# надо сравнить два файла xls файла по колонкам с фио
# за главную информацию взять информацию из файла 1С
# и найти те строки, которых нет в 1С

import openpyxl
import openpyxl.styles

def str_contains_digit(sequence):
    return any(character.isdigit() for character in sequence)

# файлы данных
file_staff_1c = 'staff_1c.xlsx'
file_staff_aduser = 'staff_aduser.xlsx'

# подключаюсь к файлу
wb_1c = openpyxl.load_workbook(file_staff_1c)
wb_1c_s = wb_1c.active

# ищу минимальные и максимальные места, где есть данные
wb_1c_s_row_begin = 1  # wb_1c_s.min_row
wb_1c_s_row_end = wb_1c_s.max_row
wb_1c_s_col_begin = 1  # wb_1c_s.min_column
wb_1c_s_col_end = wb_1c_s.max_column

# читаю данные из файла
for fio_list in wb_1c_s.iter_cols(min_col=wb_1c_s_col_begin, max_col=wb_1c_s_col_end,
                                  min_row=wb_1c_s_row_begin, max_row=wb_1c_s_row_end, values_only=True):
    pass
wb_1c.close()

# создаётся множество для хранения уникальных значений ФИО для сравнения ниже
set_fio_1c = set()
for fio in fio_list:
    if not str_contains_digit(str(fio)):
        uniq_fio = ''.join(str(fio).strip().lower().split())
        set_fio_1c.add(uniq_fio)

# подключаюсь к файлу
wb_aduser = openpyxl.load_workbook(file_staff_aduser)
wb_aduser_s = wb_aduser.active

# ищу минимальные и максимальные места, где есть данные
wb_aduser_s_row_begin = 1  # wb_aduser_s.min_row
wb_aduser_s_row_end = wb_aduser_s.max_row
wb_aduser_s_col_begin = 1  # wb_aduser_s.min_column
wb_aduser_s_col_end = wb_aduser_s.max_column

# читаю данные из файла
for fio_list in wb_aduser_s.iter_cols(min_col=wb_aduser_s_col_begin, max_col=wb_aduser_s_col_end,
                                      min_row=wb_aduser_s_row_begin, max_row=wb_aduser_s_row_end, values_only=True):
    pass

# создаётся множество для хранения уникальных значений ФИО для сравнения ниже
set_fio_aduser = set()
for fio in fio_list:
    if not str_contains_digit(str(fio)):
        uniq_fio = ''.join(str(fio).strip().lower().split())
        set_fio_aduser.add(uniq_fio)

# получаю множество тех, кто есть в АДЪЮЗЕРЕ и кого нет в 1С
set_fio_for_dismiss = set_fio_aduser - set_fio_1c

# читаю файл и заливаю колонку, которая есть в set_fio_for_dismiss цветом
style_red = openpyxl.styles.NamedStyle(name='style_red')
style_red.fill = openpyxl.styles.PatternFill('solid', fgColor='00FF0000')  # красный
for fio_list in wb_aduser_s.iter_rows():
    uniq_fio = ''.join(str(fio_list[0].value).strip().lower().split())
    if uniq_fio in set_fio_for_dismiss:
        wb_aduser_s.cell(fio_list[0].row, fio_list[0].column).style = style_red

# сохраняю и закрываю файл
wb_aduser.save(file_staff_aduser)
wb_aduser.close()
